from dotenv import load_dotenv
load_dotenv()
from openai import OpenAI
import os
import re, json

client = OpenAI(
    api_key=os.getenv("MISTRAL_API_KEY"),
    base_url="https://api.mistral.ai/v1"
)

# client = OpenAI(
#     api_key="FOUNDATION_API_KEY",
#     base_url="https://fmgateway.proxem.dsone.3ds.com/v1"
# )

ALL_HEURISTICS = [
    "request_adequacy",
    "transparency_of_reasoning",
    "contextual_relevance",
    "human_controllability",
    "cognitive_load_reduction",
    "reliability_and_anticipation",
    "task_segmentation",
    "interface_and_3d_model_relationship",
    "interoperability",
    "consistency_over_time",
]

def call_main_agent(prompt):
    response = client.chat.completions.create(
        # model="openai/gpt-oss-120b",
        model="mistral-large-latest",
        messages=[
            {"role": "system", "content": "You are an expert assistant."},
            {"role": "user", "content": prompt}
        ]
    )
    return response.choices[0].message.content

def parse_conversation(text):
    parts = re.split(r'\d{1,2}:\d{2}(?:\s*[AP]M)?', text)
    parts = [p.strip() for p in parts if p.strip()]
    if len(parts) < 2:
        return None
    exchanges = []
    for i, msg in enumerate(parts):
        role = "USER" if i % 2 == 0 else "LEO"
        exchanges.append({"role": role, "index": i + 1, "message": msg.strip()})
    return exchanges

def format_conversation_for_prompt(exchanges):
    lines = []
    exchange_num = 0
    for i in range(0, len(exchanges) - 1, 2):
        exchange_num += 1
        user_msg = exchanges[i]["message"] if i < len(exchanges) else ""
        leo_msg = exchanges[i+1]["message"] if i+1 < len(exchanges) else ""
        lines.append(f"--- EXCHANGE {exchange_num} ---")
        lines.append(f"[USER]: {user_msg}")
        lines.append(f"[LEO]: {leo_msg}")
        lines.append("")
    return "\n".join(lines)

def analyze_interface_image(image_b64):
    try:
        response = client.chat.completions.create(
            model="pixtral-12b-2409",
            messages=[
                {
                    "role": "user",
                    "content": [
                        {
                            "type": "image_url",
                            "image_url": "data:image/png;base64," + (image_b64[0] if isinstance(image_b64, list) else image_b64)
                        },
                        {
                            "type": "text",
                            "text": "You are a UX researcher analyzing a screenshot from a CAD/simulation software interface (Dassault Systemes SIMULIA / 3DEXPERIENCE). Analyze this screenshot and describe: 1. What is visible in the 3D view (objects, selections, highlights, mesh, annotations)? 2. Is there any visual feedback from the AI agent (LEO/AURA/MARIE) on the 3D model (highlighted zones, color changes, annotations, glyphs)? 3. Does the interface state seem to reflect an action performed by the AI agent? 4. What is the overall state of the interface (which panels are open, what is selected)? Be specific and concise."
                        }
                    ]
                }
            ],
            max_tokens=500
        )
        return response.choices[0].message.content
    except Exception as e:
        return "Image analysis unavailable: " + str(e)

def build_json_template(active_heuristics, has_image=False, mode="single"):
    """Build the JSON output template dynamically based on active heuristics."""
    always_na = {
        "interoperability": "Not applicable: no external data required.",
        "consistency_over_time": "Not applicable: single exchange." if mode == "single" else None,
        "interface_and_3d_model_relationship": None,
    }

    parts = {}
    for key in ALL_HEURISTICS:
        if active_heuristics and key not in active_heuristics:
            # Not selected in Focus mode → N/A
            parts[key] = f'"{key}": {{"score": null, "applicable": false, "justification": "Not evaluated in Focus mode.", "improvement_advice": "N/A"}}'
        elif key == "consistency_over_time" and mode == "single":
            parts[key] = f'"{key}": {{"score": null, "applicable": false, "justification": "Not applicable: single exchange.", "improvement_advice": "N/A"}}'
        elif key == "interoperability":
            parts[key] = f'"{key}": {{"score": null, "applicable": false, "justification": "Not applicable: no external data required.", "improvement_advice": "N/A"}}'
        elif key == "interface_and_3d_model_relationship":
            if has_image:
                parts[key] = f'"{key}": {{"score": 0, "applicable": true, "justification": "Based on the screenshot analysis provided.", "improvement_advice": "..."}}'
            else:
                parts[key] = f'"{key}": {{"score": null, "applicable": false, "justification": "Not applicable: no screenshot or 3D context provided.", "improvement_advice": "N/A"}}'
        else:
            parts[key] = f'"{key}": {{"score": 0, "applicable": true, "justification": "...", "improvement_advice": "..."}}'

    # First key gets the full example
    first_active = next((k for k in ALL_HEURISTICS if active_heuristics is None or k in active_heuristics), None)
    if first_active and first_active not in ["interoperability", "consistency_over_time", "interface_and_3d_model_relationship"]:
        parts[first_active] = f'"{first_active}": {{"score": 0, "applicable": true, "justification": "2-3 sentences, analytical, SIMULIA context, no quotes, no technical data.", "improvement_advice": "[HIGH] Title — \\"short quote\\" → UX problem.\\n[MEDIUM] Title — \\"short quote\\" → UX gap.\\n[LOW] Title — \\"quote\\" → Minor UX polish."}}'

    lines = [f"    {v}" for v in parts.values()]
    return '{{\n  "evaluation": {{\n' + ',\n'.join(lines) + '\n  }},\n  "global_improvement_suggestions": ["suggestion 1", "suggestion 2", "suggestion 3"]\n}}'

def build_focus_instruction(active_heuristics):
    if not active_heuristics:
        return ""
    names = {
        "request_adequacy": "Request Adequacy",
        "transparency_of_reasoning": "Transparency of Reasoning",
        "contextual_relevance": "Contextual Relevance",
        "human_controllability": "Human Controllability",
        "cognitive_load_reduction": "Cognitive Load Reduction",
        "reliability_and_anticipation": "Reliability & Anticipation",
        "task_segmentation": "Task Segmentation",
        "interface_and_3d_model_relationship": "Interface & 3D Model Relationship",
        "interoperability": "Interoperability",
        "consistency_over_time": "Consistency Over Time",
    }
    selected = [names[k] for k in active_heuristics if k in names]
    skipped = [names[k] for k in ALL_HEURISTICS if k not in active_heuristics and k in names]
    instruction = f"\nFOCUS MODE — Evaluate ONLY these heuristics: {', '.join(selected)}.\n"
    if skipped:
        instruction += f"Set these as NOT EVALUATED (score: null, applicable: false): {', '.join(skipped)}.\n"
    return instruction

SIMULIA_CONTEXT = """
EVALUATION CONTEXT — READ CAREFULLY:
You are evaluating a Virtual Companion (AURA, LEO, or MARIE) — an AI agent embedded inside Dassault Systèmes 3DEXPERIENCE platform (SIMULIA, CATIA, ENOVIA).

These companions assist engineers and designers in complex industrial workflows:
- Structural simulation (FEA), fluid dynamics (CFD), material selection, mesh refinement
- CAD operations, assembly design, tolerance analysis
- PLM data management, BOM, configuration management

TYPICAL USER PROFILES:
- Senior engineers (10+ years): expect precise technical answers, direct recommendations, no hand-holding
- Junior engineers / students: may need more guidance, clearer explanations, step-by-step breakdowns
- A good companion DETECTS the user's level from how they phrase their question and ADAPTS accordingly.

YOUR ROLE AS EVALUATOR:
You are a UX researcher — NOT an engineer. You evaluate INTERACTION DESIGN quality only.
You assess: response structure, user level adaptation, transparency of reasoning, controllability, cognitive load, task segmentation.
You do NOT assess technical accuracy. You do NOT suggest technical improvements.
"""

ADVICE_FORMAT_INSTRUCTION = """
CRITICAL — improvement_advice FORMAT:
You MUST format improvement_advice as a structured list of UX priorities. Each item MUST follow this exact format:

[HIGH] Short title — "exact short quote from agent response (max 8 words)" → UX problem identified: explain what interaction pattern or behaviour is missing from a UX perspective.
[MEDIUM] Short title — "exact short quote or description of absence" → UX gap explained: what should the agent have done differently in terms of interaction design.
[LOW] Short title — "quote or description of absence" → Minor UX polish: what small improvement would enhance the experience.

STRICT RULES:
- Maximum 3 bullets per criterion
- Each bullet MUST include a SHORT direct quote from the agent response (max 8 words) OR clearly state what is absent
- The arrow → explains the UX PROBLEM ONLY — describe what interaction pattern or UX behaviour is missing
- ABSOLUTELY NO technical content after the arrow — no material names, no numerical values, no software-specific references, no simulation parameters
- Stay 100% in UX territory: user level detection, response structure, information architecture, interaction patterns, controllability, transparency
- Good examples: "agent should detect user expertise before responding" / "response lacks a proposed next action" / "no confirmation step before proceeding"
- Bad examples: "agent should cite T700 carbon fiber" / "agent should reference CATIA material library" / "agent should use E=135 GPa"
- Keep each bullet under 2 lines
- Prioritize: HIGH = blocks or significantly degrades user task, MEDIUM = degrades experience, LOW = polish
- If score is 4 or 5, still provide 1 LOW item for continuous improvement
- ⚠️ FINAL WARNING: The evaluator is a UX researcher, NOT an engineer. Any numerical value, material name, or technical specification in improvement_advice is a CRITICAL ERROR. Write ONLY about interaction design.
"""

JUSTIFICATION_INSTRUCTION = """
CRITICAL — justification FORMAT:
Write 2-3 sentences maximum. Explain WHY this score was given in plain analytical language.
- Reference the SIMULIA/industrial context (e.g. "An engineer working in SIMULIA would expect...")
- Mention user level detection if relevant
- NO citations or quotes in justification — save quotes for improvement_advice
- Be direct and specific, not generic
- NO technical data fabrication — describe UX problems, not engineering solutions
"""

def evaluate_response(prompt, response_text, language="en", mode="single", conversation_raw="", image_b64=None, user_comment=None, active_heuristics=None):
    if language == "fr":
        lang_instruction = "Reponds UNIQUEMENT en francais. Tous les champs du JSON doivent etre rediges en francais. Les labels [HIGH]/[MEDIUM]/[LOW] restent en anglais."
    else:
        lang_instruction = "Respond ONLY in English. All JSON fields must be written in English."

    focus_instruction = build_focus_instruction(active_heuristics)
    has_image = bool(image_b64)

    if mode == "single":
        comment_text = ""
        if user_comment:
            comment_text = "\n\nUSER CONTEXT NOTE: " + str(user_comment)

        image_analysis_text = ""
        if image_b64:
            image_analysis = analyze_interface_image(image_b64)
            image_analysis_text = "\n\nINTERFACE SCREENSHOT ANALYSIS (use for Criterion 8):\n" + image_analysis + "\n\nCriterion 8 IS NOW APPLICABLE."

        json_template = build_json_template(active_heuristics, has_image=has_image, mode="single")

        evaluation_prompt = f"""
{SIMULIA_CONTEXT}

LANGUAGE INSTRUCTION: {lang_instruction}

{ADVICE_FORMAT_INSTRUCTION}

{JUSTIFICATION_INSTRUCTION}
{focus_instruction}

You are evaluating a SINGLE EXCHANGE between a user and a Virtual Companion (LEO/AURA/MARIE).

HEURISTIC EVALUATION FRAMEWORK
================================

CRITERION 1 - Request Adequacy
Did the agent understand the request and respond in the right format?
0: Completely misses the request.
1: Wrong format or heavily partial — key parts ignored.
2: Right topic but wrong format or missing significant parts.
3: Main request covered with correct format, minor gaps.
4: Well-structured, full coverage, immediately readable.
5: Perfect — every part answered in ideal format, nothing missing.

CRITERION 2 - Transparency of Reasoning ("Why")
Does the agent show WHY it recommends what it recommends?
0: Pure black box — conclusion with zero explanation.
1: Vague generic phrase — no actual reasoning.
2: Some reasoning but key assumptions hidden.
3: Main reasoning visible, some gaps — partially traceable.
4: Clear reasoning chain, minor gaps.
5: Full reasoning, sources/assumptions cited, user invited to verify.

CRITERION 3 - Contextual Relevance ("Where/Who")
Does the agent adapt to the user's role, expertise level, and workflow step?
0: Completely generic — no industrial context whatsoever.
1: Mentions domain but ignores user level or workflow step.
2: Right domain but wrong level (too basic for expert, too complex for junior).
3: Reasonable adaptation, minor vocabulary or level mismatches.
4: Detects user level, adapts vocabulary, stays in workflow context.
5: Detects level from phrasing, adapts depth, offers to adjust, references specific workflow step.

CRITERION 4 - Human Controllability
Does the agent preserve user agency?
0: Irreversible actions implied with no warning.
1: No mechanism to override or cancel.
2: Vague implication that user could modify.
3: Acknowledges user control in principle.
4: Actively invites validation before proceeding.
5: Every suggestion framed as optional, explains how to modify/cancel.

CRITERION 5 - Cognitive Load Reduction
Does the agent make the task easier?
0: Overwhelming wall of text OR completely empty.
1: Right content but chaotic structure.
2: Understandable but wrong size for the task.
3: Reasonable but could be tighter.
4: Well-calibrated length and structure.
5: Perfect length, clear formatting, summary offered when needed.

CRITERION 6 - Reliability & Anticipation ("Guardrails")
Does the agent handle uncertainty and prevent errors?
0: Confidently answers ambiguous questions, no error prevention.
1: Guesses without flagging uncertainty.
2: Generic disclaimer only — no targeted uncertainty management.
3: Identifies ambiguity, asks one clarifying question.
4: Identifies + asks + proposes alternative + mentions downstream risk.
5: Flags all uncertainty, targeted questions, multiple alternatives, anticipates effects.

CRITERION 7 - Task Segmentation ("How")
Does the agent break complex tasks into clear steps?
0: No decomposition — dumps all information as one block.
1: Lists items but unordered or incomplete.
2: Partial sequence — steps present but dependencies missing.
3: Logical sequence with some gaps.
4: Clear ordered steps, mostly complete.
5: Complete sequence — prerequisites checked, dependencies explicit.

CRITERION 8 - Interface & 3D Model Relationship
NOT APPLICABLE if no 3D model interaction and no screenshot provided.
0: No 3D awareness.
1: Vague 3D reference.
2: Acknowledges 3D object but no interface action.
3: References specific objects, proposes one interface action.
4: Partial 3D interaction.
5: Full sync — highlights, annotations, interface and dialogue aligned.

CRITERION 9 - Interoperability (Data Access)
NOT APPLICABLE if no external data needed.

CRITERION 10 - Consistency Over Time ("Memory")
NOT APPLICABLE for a single exchange.

USER PROMPT: {prompt}
AGENT RESPONSE: {response_text}
{comment_text}
{image_analysis_text}

OUTPUT — STRICTLY VALID JSON — NO EXTRA TEXT:
{json_template}
STRICT RULES: justification = 2-3 sentences max, analytical, SIMULIA context, NO quotes, NO technical data. improvement_advice = [HIGH]/[MEDIUM]/[LOW] bullets, arrow explains UX gap NOT a technical rewrite, NEVER invent numerical values. Respond ONLY with JSON.
"""

    else:
        exchanges = parse_conversation(conversation_raw)
        if not exchanges:
            return evaluate_response(prompt, conversation_raw, language, "single", "", active_heuristics=active_heuristics)

        conversation_formatted = format_conversation_for_prompt(exchanges)
        n_exchanges = len(exchanges) // 2
        json_template = build_json_template(active_heuristics, has_image=has_image, mode="multi")

        evaluation_prompt = f"""
{SIMULIA_CONTEXT}

LANGUAGE INSTRUCTION: {lang_instruction}

{ADVICE_FORMAT_INSTRUCTION}

{JUSTIFICATION_INSTRUCTION}
{focus_instruction}

You are evaluating a FULL CONVERSATION between a user and a Virtual Companion (LEO/AURA/MARIE). The conversation has {n_exchanges} exchanges.

For each criterion:
- Give a GLOBAL score (0-5) reflecting overall quality across ALL exchanges
- justification: 2-3 sentences analytical, reference specific exchange numbers, SIMULIA context, NO quotes, NO technical data
- improvement_advice: [HIGH]/[MEDIUM]/[LOW] bullets — arrow explains UX gap, cite exchange numbers, NEVER invent technical data

CONVERSATION TO EVALUATE:
{conversation_formatted}

HEURISTIC EVALUATION FRAMEWORK
================================

CRITERION 1 - Request Adequacy
0: Misses requests consistently. 1: Wrong format in most exchanges. 2: Right topic, significant gaps. 3: Main requests covered, minor gaps. 4: Well-structured throughout. 5: Every request perfectly addressed.

CRITERION 2 - Transparency of Reasoning
0: Black box throughout. 1: Vague generic explanations. 2: Partial reasoning in some exchanges. 3: Adequate reasoning, some gaps. 4: Clear reasoning throughout. 5: Full transparency, sources cited.

CRITERION 3 - Contextual Relevance
0: Completely generic throughout. 1: Minimal adaptation. 2: Right domain, wrong level. 3: Reasonable adaptation overall. 4: Tracks user level across exchanges. 5: Perfect adaptation, offers level adjustment.

CRITERION 4 - Human Controllability
0: Irreversible actions throughout. 1: No override mechanisms. 2: Vague implication of control. 3: Acknowledges control in principle. 4: Actively invites validation. 5: Full controllability throughout.

CRITERION 5 - Cognitive Load Reduction
0: Overwhelming or empty throughout. 1: Poor structure overall. 2: Wrong size in most responses. 3: Reasonable but could be tighter. 4: Well-calibrated throughout. 5: Optimal across all exchanges.

CRITERION 6 - Reliability & Anticipation
0: Confident on ambiguous requests throughout. 1: Guesses without flagging. 2: Generic disclaimers only. 3: Identifies some ambiguities. 4: Good uncertainty management. 5: Exemplary throughout.

CRITERION 7 - Task Segmentation
0: No decomposition throughout. 1: Unordered steps. 2: Partial sequences. 3: Logical sequences, some gaps. 4: Clear ordered sequences. 5: Complete sequences with dependencies.

CRITERION 8 - Interface & 3D Model Relationship
NOT APPLICABLE if no 3D interaction in conversation.

CRITERION 9 - Interoperability
NOT APPLICABLE if no external data needed.

CRITERION 10 - Consistency Over Time ("Memory") — KEY CRITERION
APPLICABLE for multi-exchange conversations.
0: No memory, fresh start each exchange. 1: Vague inconsistent references. 2: Recalls some info inconsistently. 3: Reasonable memory, misses contradictions. 4: Good memory, references earlier decisions. 5: Perfect memory, proactively applies context.

OUTPUT — STRICTLY VALID JSON — NO EXTRA TEXT:
{json_template}
STRICT RULES: justification = 2-3 sentences max, cite exchange numbers, SIMULIA context, NO quotes, NO technical data. improvement_advice = [HIGH]/[MEDIUM]/[LOW] format, arrow = UX gap NOT technical rewrite, NEVER invent numerical values. Respond ONLY with JSON.
- ⚠️ FINAL WARNING: The evaluator is a UX researcher, NOT an engineer. Any numerical value, material name, or technical specification in improvement_advice is a CRITICAL ERROR. Write ONLY about interaction design.
"""

    evaluation = client.chat.completions.create(
        # model="openai/gpt-oss-120b",
        model="mistral-large-latest",
        messages=[
            {"role": "system", "content": "You are a senior UX researcher evaluating AI agents in industrial software. You respond only in valid JSON."},
            {"role": "user", "content": evaluation_prompt}
        ]
    ).choices[0].message.content

    match = re.search(r'\{.*\}', evaluation, re.DOTALL)
    if match:
        evaluation_json = match.group(0)
        try:
            data = json.loads(evaluation_json)
            for c in data["evaluation"]:
                criterion = data["evaluation"][c]
                if criterion.get("applicable") == False:
                    continue
                score = criterion.get("score")
                if score is None or not isinstance(score, (int, float)):
                    continue
                data["evaluation"][c]["score"] = max(0, min(5, score))
            return json.dumps(data)
        except json.JSONDecodeError:
            return evaluation_json
    else:
        return evaluation


def generate_export_xlsx(result: dict, title: str = "Exchange 1") -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    import io

    FONT_NAME = "Aptos Narrow"
    CRITERION_COLORS = ["B8D9F5","AADEC8","B8E3A0","E0EE9A","F7E08A","F9C48A","F4A89A","F0A0C0","D4AEE8","B0BAEE"]
    LETTERS = ["A","B","C","D","E","F","G","H","I","J"]
    CRITERION_LABELS = [
        "A. Request adequacy","B. Reasoning transparency","C. Contextual relevance",
        "D. Human controllability","E. Cognitive load reduction","F. Reliability & anticipation",
        "G. Action segmentation","H. Bridge to interface and 3D model",
        "I. Interoperability","J. Consistency over time",
    ]

    def fill(hex_color): return PatternFill("solid", fgColor=hex_color)
    def thin_border(left=True, right=True, top=True, bottom=True):
        thin = Side(border_style="thin"); none = Side(border_style=None)
        return Border(left=thin if left else none, right=thin if right else none, top=thin if top else none, bottom=thin if bottom else none)
    def font(size=11, bold=False): return Font(name=FONT_NAME, size=size, bold=bold)

    wb = Workbook(); ws = wb.active; ws.title = "Sheet1"

    ws.column_dimensions["B"].width = 45
    for col, w in zip("CDEFGHIJKLMN", [4,4,4,4,4,4,4,4,4,4,7.1,8.7]):
        ws.column_dimensions[col].width = w

    ws.merge_cells("C2:N3"); ws["C2"] = title
    ws["C2"].font = Font(name=FONT_NAME, size=18, bold=True)
    ws["C2"].fill = fill("FFD7C8")
    ws["C2"].alignment = Alignment(horizontal="center", vertical="middle")
    ws.row_dimensions[2].height = 21; ws.row_dimensions[3].height = 21

    ws.merge_cells("C4:N4")
    ws["C4"] = "Grades for the 10 Criteria"
    ws["C4"].font = Font(name=FONT_NAME, size=18, bold=True, color="FFFFFFFF")
    ws["C4"].fill = fill("4BACC6")
    ws["C4"].alignment = Alignment(horizontal="center", vertical="middle")
    ws.row_dimensions[4].height = 24

    ws.row_dimensions[5].height = 21
    for i, (letter, color) in enumerate(zip(LETTERS, CRITERION_COLORS)):
        col = get_column_letter(3 + i); cell = ws[f"{col}5"]
        cell.value = letter; cell.fill = fill(color)
        cell.font = font(11); cell.alignment = Alignment(horizontal="center"); cell.border = thin_border()
    ws["M5"] = "Avg"; ws["M5"].font = font(11); ws["M5"].alignment = Alignment(horizontal="center")
    ws["N5"] = "Sum"; ws["N5"].font = font(11); ws["N5"].alignment = Alignment(horizontal="center")

    ws.row_dimensions[6].height = 21
    criteria = result.get("criteria", [])
    for i in range(10):
        col = get_column_letter(3 + i); cell = ws[f"{col}6"]
        if i < len(criteria):
            score = criteria[i].get("score")
            cell.value = score if score is not None else "N/A"
        cell.font = font(11); cell.alignment = Alignment(horizontal="center")
    ws["M6"] = "=AVERAGE(C6:L6)"; ws["M6"].font = font(11); ws["M6"].alignment = Alignment(horizontal="center")
    ws["N6"] = "=SUM(C6:L6)"; ws["N6"].font = font(11); ws["N6"].alignment = Alignment(horizontal="center")

    ws.row_dimensions[7].height = 24
    ws["B7"] = "10 Criteria"
    ws["B7"].font = Font(name=FONT_NAME, size=11, bold=True, color="FFFFFFFF")
    ws["B7"].fill = fill("4BACC6")
    ws["B7"].alignment = Alignment(horizontal="left", vertical="middle")
    ws.merge_cells("C7:N7")
    ws["C7"] = "Detailed evaluations"
    ws["C7"].font = Font(name=FONT_NAME, size=18, bold=True, color="FFFFFFFF")
    ws["C7"].fill = fill("4BACC6")
    ws["C7"].alignment = Alignment(horizontal="center", vertical="middle")

    for i in range(10):
        start_row = 8 + i * 2; color = CRITERION_COLORS[i]
        ws.merge_cells(f"B{start_row}:B{start_row+1}")
        b = ws[f"B{start_row}"]
        b.value = CRITERION_LABELS[i]; b.fill = fill(color); b.font = font(14)
        b.alignment = Alignment(horizontal="left", vertical="center"); b.border = thin_border()

        cdata = criteria[i] if i < len(criteria) else {}
        for j, (label, key) in enumerate([("Why this score", "justification"), ("How to improve", "advice")]):
            row = start_row + j; value = cdata.get(key, "") or ""
            ws.merge_cells(f"C{row}:N{row}")
            c = ws[f"C{row}"]
            c.value = f"{label}: {value}" if value else label
            c.font = font(11); c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            c.border = thin_border(left=True, right=False, top=j==0, bottom=j==1)
            charsPerLine = 90
            lines = max(1, len(str(value)) // charsPerLine + 1) if value else 1
            ws.row_dimensions[row].height = max(20, lines * 15)

    ws.row_dimensions[40].height = 37.5
    ws["B40"] = "Tips: Use ALT+H+O+I to autofit columns width"; ws["B40"].font = font(10)

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf.read()